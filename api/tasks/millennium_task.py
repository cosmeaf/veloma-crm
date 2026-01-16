from celery import shared_task
import tempfile
import uuid
import os
import openpyxl

from django.conf import settings
from api.utils.millennium.extractor_pdf import MillenniumExtractor


@shared_task
def process_millennium_pdf(pdf_bytes):
    temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_pdf.write(pdf_bytes)
    temp_pdf.close()

    extractor = MillenniumExtractor(temp_pdf.name)
    movimentos = extractor.extract()

    template_path = os.path.join(
        settings.BASE_DIR,
        "api/templates/excel_templates/template_importacao.xlsx"
    )

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    row = 2
    for mov in movimentos:
        ws.cell(row=row, column=1).value = mov["data_lanc"]
        ws.cell(row=row, column=2).value = mov["data_valor"]
        ws.cell(row=row, column=3).value = mov["descricao"]
        ws.cell(row=row, column=4).value = mov["debito"]
        ws.cell(row=row, column=5).value = mov["credito"]
        ws.cell(row=row, column=6).value = mov["saldo"]
        row += 1

    output_dir = os.path.join(settings.MEDIA_ROOT, "exports")
    os.makedirs(output_dir, exist_ok=True)

    file_path = os.path.join(output_dir, f"millennium_{uuid.uuid4()}.xlsx")
    wb.save(file_path)

    return file_path
