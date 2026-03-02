# =========================================================================================
#  MILLENIUM BCP PDF EXTRACTOR - COMPATÍVEL COM TEMPLATE TOCONLINE
# =========================================================================================

import uuid
import re
from dataclasses import dataclass
from datetime import date
from typing import List, Optional
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from django.conf import settings

from converter.models.millenium_model import MilleniumFile


# =========================================================================================
# CONSTANTES
# =========================================================================================

MEDIA_ROOT = Path(settings.MEDIA_ROOT)
TEMPLATE_DIR = MEDIA_ROOT / "converter/upload"
JOB_TMP_DIR = Path("/var/tmp/veloma_pdf_jobs")

MONEY_RX = re.compile(r"\d{1,3}(?:\s\d{3})*(?:[.,]\d{2})")
LINE_RX = re.compile(r"^(\d{1,2}\.\d{2})\s+(\d{1,2}\.\d{2})\s+(.+)$")
PERIOD_RX = re.compile(
    r"(\d{4})/(\d{2})/(\d{2})\s+A\s+(\d{4})/(\d{2})/(\d{2})"
)


# =========================================================================================

@dataclass
class Movement:
    data_mov: date
    data_valor: date
    descricao: str
    valor: float


# =========================================================================================

class MilleniumExtractor:

    def __init__(self, bank_name: str = "millenium", pdf_path: Optional[str] = None):

        self.bank_name = bank_name.lower().strip()
        self.job_id = uuid.uuid4()

        self.job_dir = JOB_TMP_DIR / str(self.job_id)
        self.job_dir.mkdir(parents=True, exist_ok=True)

        if pdf_path:
            self.pdf_path = Path(pdf_path)
        else:
            latest = MilleniumFile.objects.order_by("-created_at").first()
            if not latest:
                raise ValueError("NO_PDF_UPLOADED")
            self.pdf_path = Path(latest.file.path)

    # =====================================================================================
    # EXECUÇÃO PRINCIPAL
    # =====================================================================================

    def run(self) -> Path:

        if not self.pdf_path.exists():
            raise FileNotFoundError(f"PDF_NOT_FOUND: {self.pdf_path}")

        template_path = self._get_template()
        text = self._extract_text()

        period = self._extract_period(text)
        year = period["start_year"]

        saldo_inicial = self._extract_saldo(text, "SALDO INICIAL")
        saldo_final = self._extract_saldo(text, "SALDO FINAL")

        movements = self._extract_movements(
            text=text,
            year=year,
            saldo_inicial=saldo_inicial
        )

        return self._fill_excel(
            template_path=template_path,
            saldo_inicial=saldo_inicial,
            saldo_final=saldo_final,
            movements=movements
        )

    # =====================================================================================
    # TEMPLATE
    # =====================================================================================

    def _get_template(self) -> Path:

        if not TEMPLATE_DIR.exists():
            raise FileNotFoundError(f"TEMPLATE_DIR_NOT_FOUND: {TEMPLATE_DIR}")

        xlsx_files = sorted(
            TEMPLATE_DIR.rglob("*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True
        )

        if not xlsx_files:
            raise FileNotFoundError("NO_TEMPLATE_FOUND")

        return xlsx_files[0]

    # =====================================================================================
    # EXTRAÇÃO PDF
    # =====================================================================================

    def _extract_text(self) -> str:

        content = ""

        with pdfplumber.open(self.pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""

                if "NEWS EMPRESAS" in page_text:
                    continue

                content += page_text + "\n"

        return re.sub(r"[ \t]+", " ", content)

    # =====================================================================================
    # PERÍODO
    # =====================================================================================

    def _extract_period(self, text: str):

        m = PERIOD_RX.search(text)
        if not m:
            raise ValueError("PERIOD_NOT_FOUND")

        return {
            "start_year": int(m.group(1))
        }

    # =====================================================================================
    # SALDOS
    # =====================================================================================

    def _extract_saldo(self, text: str, keyword: str) -> float:

        m = re.search(rf"{keyword}\s+(.+)", text)
        if not m:
            raise ValueError(f"{keyword}_NOT_FOUND")

        mm = MONEY_RX.search(m.group(1))
        if not mm:
            raise ValueError(f"{keyword}_NUMBER_NOT_FOUND")

        return self._to_float(mm.group(0))

    # =====================================================================================
    # MOVIMENTOS
    # =====================================================================================

    def _extract_movements(
        self,
        text: str,
        year: int,
        saldo_inicial: float
    ) -> List[Movement]:

        movements: List[Movement] = []
        saldo_anterior = saldo_inicial

        for line in text.splitlines():

            m = LINE_RX.match(line.strip())
            if not m:
                continue

            dm, dv, rest = m.groups()
            amounts = MONEY_RX.findall(rest)

            if len(amounts) < 2:
                continue

            saldo_atual = self._to_float(amounts[-1])
            valor = round(saldo_atual - saldo_anterior, 2)
            saldo_anterior = saldo_atual

            descricao = MONEY_RX.sub("", rest).strip()

            movements.append(
                Movement(
                    data_mov=self._parse_mmdd(dm, year),
                    data_valor=self._parse_mmdd(dv, year),
                    descricao=descricao,
                    valor=valor,
                )
            )

        return movements

    # =====================================================================================
    # EXPORTAÇÃO EXCEL (FORMATO FIXO TOCONLINE)
    # =====================================================================================

    def _fill_excel(
        self,
        template_path: Path,
        saldo_inicial: float,
        saldo_final: float,
        movements: List[Movement]
    ) -> Path:

        filename = f"{self.bank_name}.{self.job_id}.xlsx"
        output_path = self.job_dir / filename

        try:
            wb = load_workbook(template_path)
        except Exception as e:
            raise ValueError(f"TEMPLATE_INVALID: {e}")

        ws = wb.active

        # -------------------------------------------------------------------------
        # NÃO ALTERAR ESTRUTURA DO TEMPLATE
        # -------------------------------------------------------------------------

        ws["D3"].value = saldo_inicial
        ws["D5"].value = saldo_final

        ws["D3"].number_format = "#,##0.00"
        ws["D5"].number_format = "#,##0.00"

        start_row = 8

        # Limpar linhas antigas sem tocar no header
        for row in range(start_row, ws.max_row + 1):
            ws.cell(row, 1).value = None
            ws.cell(row, 2).value = None
            ws.cell(row, 3).value = None
            ws.cell(row, 4).value = None

        # Escrever movimentos
        for idx, mv in enumerate(movements):
            r = start_row + idx

            ws.cell(r, 1).value = mv.data_mov
            ws.cell(r, 1).number_format = "DD/MM/YYYY"

            ws.cell(r, 2).value = mv.data_valor
            ws.cell(r, 2).number_format = "DD/MM/YYYY"

            ws.cell(r, 3).value = mv.descricao

            ws.cell(r, 4).value = float(mv.valor)
            ws.cell(r, 4).number_format = "#,##0.00"

        try:
            wb.save(output_path)
        except Exception as e:
            raise RuntimeError(f"FAILED_TO_SAVE_XLSX: {e}")
        finally:
            wb.close()

        if not output_path.exists() or output_path.stat().st_size == 0:
            raise RuntimeError("OUTPUT_FILE_NOT_CREATED")

        return output_path

    # =====================================================================================

    @staticmethod
    def _to_float(value: str) -> float:
        return float(value.replace(" ", "").replace(",", "."))

    @staticmethod
    def _parse_mmdd(token: str, year: int) -> date:
        mm, dd = token.split(".")
        return date(year, int(mm), int(dd))